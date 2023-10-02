# Copyright 2023 ©Mohamed Rizad
# 250923 @riz4d
import openpyxl
from modules.font import Font
from modules.align import Alignment

headers = []
output_file_id = input("Enter Output Filename : ")
num_of_heads = int(input("Enter How many heads are you using ? "))
for i in range(num_of_heads):
    i+=1
    head_id=input(f"Enter Head {i} Name : ")
    headers.append(head_id)
    

workbook = openpyxl.Workbook()
sheet = workbook.active

for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

for col_num in range(1, num_of_heads + 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.alignment = Alignment(horizontal='center') 
lettr = ord('a')
def enter_data(chacode,datalst):
    letterr = ord('a')
    next_row = sheet.max_row + 1
    for i in range(len(datalst)):
        charcode = chr(letterr).upper()
        sheet[f'{charcode}{next_row}'] = datalst[i]
        letterr+=1
def datadbm(*data):
    charcode = chr(data[0]).upper()
    datadb = data[1]
    enter_data(charcode,datadb)
    print(datadb)

i=0
# 250923 @riz4d
while True:
    i+=1
    dataset = []
    for i in range(len(headers)):
        dat = input(f"Enter {headers[i]} : ")
        dataset.append(dat)
    datadbm(lettr,dataset)
    lettr+=1
    workbook.save(f'{output_file_id}.xlsx')
