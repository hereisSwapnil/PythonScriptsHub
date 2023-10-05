import os
from openpyxl import load_workbook
import win32print
import win32ui

def print_excel_sheet(excel_file_path):
    try:
        # Load the Excel workbook
        wb = load_workbook(filename=excel_file_path, read_only=True)
        
        # Get the first sheet (you can modify this to print a specific sheet)
        sheet = wb.active

        # Create a temporary text file to store the sheet content
        temp_txt_file = "temp_sheet.txt"
        with open(temp_txt_file, "w") as txt_file:
            for row in sheet.iter_rows(values_only=True):
                txt_file.write("\t".join(map(str, row)) + "\n")

        # Get the default printer
        printer_name = win32print.GetDefaultPrinter()

        # Initialize the printer and document properties
        printer_handle = win32print.OpenPrinter(printer_name)
        printer_info = win32print.GetPrinter(printer_handle, 2)
        hprinter = win32print.OpenPrinter(printer_name, {"DesiredAccess": win32print.PRINTER_ACCESS_USE})

        # Print the text file to the printer
        printer_data = open(temp_txt_file, "rb").read()
        win32print.StartDocPrinter(hprinter, 1, ("Excel Sheet", None, "RAW"))
        win32print.StartPagePrinter(hprinter)
        win32print.WritePrinter(hprinter, printer_data)
        win32print.EndPagePrinter(hprinter)
        win32print.EndDocPrinter(hprinter)

        # Clean up temporary files
        os.remove(temp_txt_file)

        print("Excel sheet printed successfully.")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    excel_file_path = "Printing Excel sheet\Orders-With Nulls.xlsx"  # Replace with the path to your Excel file
    print_excel_sheet(excel_file_path)
